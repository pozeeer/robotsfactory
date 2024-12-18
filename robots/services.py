import os

from django.utils.dateparse import parse_datetime
from django.utils.timezone import make_aware
from django.core.exceptions import ValidationError
from django.conf import settings
from django.db.models import QuerySet, Count
from django.utils.timezone import now
from datetime import timedelta
from typing import Dict, List, Any
from openpyxl import Workbook

from robots.models import Robot


class RobotService:
    @staticmethod
    def create_robot(data: Dict[str, Any]) -> Robot:
        """
        Обрабатывает данные, создает запись в базе и возвращает объект робота.

        :param data: Словарь с данными о роботе (модель, версия, дата создания).
        :return: Объект модели Robot.
        :raises ValidationError: Если данные некорректны или отсутствуют.
        """
        model = data.get("model")
        version = data.get("version")
        created = parse_datetime(data.get("created"))

        if not model or not version or not created:
            raise ValidationError("Invalid data format")

        created = make_aware(created)
        robot = Robot(
            model=model,
            version=version,
            serial=f"{model}-{version}",
            created=created
        )
        robot.save()
        return robot


class ExcelReportGenerator:
    """
    Класс для генерации Excel-отчетов по производству роботов.
    """

    def __init__(self) -> None:
        """
        Инициализация объекта. Создает рабочую книгу и пути к файлам.
        """
        self.workbook = Workbook()
        self.reports_dir = os.path.join(settings.MEDIA_ROOT, "reports")
        self.file_name = "robot_report.xlsx"
        self.file_path = os.path.join(self.reports_dir, self.file_name)

    def get_last_week_data(self) -> QuerySet[Robot]:
        """
        Выбирает данные о роботах за последнюю неделю.

        :return: QuerySet с объектами модели Robot за последнюю неделю.
        :raises ValueError: Если данные отсутствуют.
        """
        current_date = now()
        last_week_date = current_date - timedelta(days=7)
        robots = Robot.objects.filter(created__gte=last_week_date)

        if not robots.exists():
            raise ValueError("Нет данных для формирования отчета.")

        return robots

    def group_data_by_model(self, robots: QuerySet[Robot]) -> Dict[str, List[Dict[str, Any]]]:
        """
        Группирует данные по моделям и версиям.

        :param robots: QuerySet с объектами модели Robot.
        :return: Словарь, где ключ — модель робота, а значение — список версий с количеством.
        """
        grouped_data = {}
        models = robots.values("model").distinct()

        for model in models:
            versions = (
                robots.filter(model=model["model"])
                .values("version")
                .annotate(count=Count("id"))
            )
            grouped_data[model["model"]] = list(versions)

        return grouped_data

    def create_sheets(self, grouped_data: Dict[str, List[Dict[str, Any]]]) -> None:
        """
        Создает страницы в Excel-файле для каждой модели.

        :param grouped_data: Словарь с данными, сгруппированными по моделям и версиям.
        """
        for model, versions in grouped_data.items():
            sheet = self.workbook.create_sheet(title=model)
            sheet.append(["Модель", "Версия", "Количество за неделю"])

            for version in versions:
                sheet.append([model, version["version"], version["count"]])

        # Удаляем стандартный пустой лист
        if "Sheet" in self.workbook.sheetnames:
            self.workbook.remove(self.workbook["Sheet"])

    def save_to_file(self) -> None:
        """
        Сохраняет файл в файловую систему.
        """
        os.makedirs(self.reports_dir, exist_ok=True)
        self.workbook.save(self.file_path)

    def generate_report(self) -> str:
        """
        Основной метод для генерации отчета.

        :return: Путь к сгенерированному файлу.
        :raises ValueError: Если данные отсутствуют.
        :raises RuntimeError: Если произошла ошибка при генерации отчета.
        """
        try:
            robots = self.get_last_week_data()
            grouped_data = self.group_data_by_model(robots)
            self.create_sheets(grouped_data)
            self.save_to_file()
            return os.path.join(settings.MEDIA_URL, "reports", self.file_name)
        except ValueError as e:
            raise e
        except Exception as e:
            raise RuntimeError(f"Ошибка при генерации отчета: {str(e)}")
